"""
Read and write tabular data from XLSX spreadsheets.

Multiple worksheets per file are supported. Each worksheet is assumed
to hold a table of data, with column headings - with perhaps a heading
or something in the first couple of rows.

See:
    https://openpyxl.readthedocs.io/
"""

import io
from pathlib import Path
from typing import Any, Dict, Iterator, Iterable, List, Optional, Tuple
import zipfile

from openpyxl import load_workbook, Workbook
from openpyxl.utils.exceptions import InvalidFileException

from animal3.utils.algorithms import rstrip_iterable
from animal3.utils.text import make_slugs

from .exceptions import ExcelReaderError


def count_values(row: Iterable[Any]) -> int:
    """
    Count number of non-none values in given iterable.

    Args:
        Iterable of values, some of which may be none.

    Returns:
        Count of 'valid' values.
    """
    count = 0
    for value in row:
        if value is not None:
            count += 1
    return count


def open_spreadsheet(path: Path, mode='r') -> Workbook:
    """
    Attempt to open spreadsheet file.

    Args:
        path:
            Path to XLSX file.
        mode:
            Open in read-only ('r') or write ('w') mode. Read-only is faster
            and safer, hence also the default.

    Raises:
        ExcelReaderError:
            If opening spreadsheet fails.

    Returns:
        openpyxl workbook.
    """
    try:
        if mode == 'r':
            workbook = load_workbook_readonly(path)
        elif mode == 'w':
            workbook = load_workbook(path, data_only=True)
        else:
            raise ExcelReaderError(f"Unknown mode: {mode!r}")
    except (InvalidFileException, KeyError, zipfile.BadZipFile):
        raise ExcelReaderError(f"Invalid file given: {path.name!r}") from None
    except (FileNotFoundError, IOError) as e:
        raise ExcelReaderError(e) from None

    return workbook


def load_workbook_readonly(path: Path) -> Workbook:
    """
    Like `openpyxl.load_workbook()`, but loads entire file into memory.

    Opening the file with `read_only` enabled boosts performance, but
    caused me intermittent python `ResourceWarning` errors when running my
    unittests under 'python -X dev'.

    As our files are all < 1MB opening copying them into a memory buffer
    is an ugly but effective workaround.

    See:
        https://stackoverflow.com/a/45387809/172949
    """
    with open(path, 'rb') as fp:
        file_bytes = io.BytesIO(fp.read())

    return load_workbook(file_bytes, data_only=True, read_only=True)


class ExcelReader:
    """
    Read spreadsheet data in faster and safer read-only mode.

    Like `csv.DictReader` it builds lists of field dictionaries, using names
    found in the header row for keys.

        >>> x = ExcelReader(path)
        >>> data = [row for row in self.reader.read()]
        >>> pprint(data[-1])


    OpenPyXL refers to an Excel spreadsheet as a workbook. A workbook contains
    at least one worksheet. The first worksheet is selected by default.

    Operations are made to only the active worksheet.
    """
    def __init__(self, path: Path):
        """
        Initialiser.

        Args:
            path:
                Path to XLSX file.
        """
        self.path = Path(path)
        self.workbook = open_spreadsheet(self.path)
        self.reset()

    def get_worksheet(self) -> str:
        """
        Return the title of the currently active worksheet.

        Returns:
            Name of active worksheet.
        """
        return self.workbook.active.title

    def get_worksheets(self) -> List[str]:
        """
        Return the title of all the available worksheets.

        Returns:
            List of worksheet names.
        """
        return self.workbook.sheetnames

    def read(
        self,
        *,
        restkey: Optional[str] = None,
        restval: Optional[Any] = None,
    ) -> Iterator[Dict[Optional[str], Any]]:
        """
        Generator over non-empty rows.

        Uses the first row as header and stops when an empty row
        is encountered.

        Note that the types of the spreadsheet cells carry through into the
        values generated here; unlike `csv.DictReader` the values are not all
        strings.

        Args:
            restkey:
                The key used to hold excess row data.
            restval:
                The value used to fill-in missing row data.

        Returns:
            Iterates over each row, producing a dictionary of values.
        """
        rows = self.workbook.active.iter_rows(values_only=True)

        # Find header
        header = self._find_header(rows)
        fieldnames = self._build_fieldnames(header)

        # Prevent data loss if keys clash
        if restkey in fieldnames:
            raise ValueError(f"Given restkey {restkey!r} clashes with fieldnames")

        # Read data
        for row in rows:
            # Stop when an empty row encountered - otherwise OpenPyXL will often
            # iterate over 1 million+ possible rows.
            if not any(row):
                break

            # Build output row
            row = rstrip_iterable(row)
            data = dict(zip(fieldnames, row))

            # Fill in any excess
            num_headings = len(fieldnames)
            num_fields = len(row)

            if num_headings < num_fields:
                data[restkey] = row[num_headings:]
            elif num_headings > num_fields:
                for key in fieldnames[num_fields:]:
                    data[key] = restval

            yield data

    def reset(self) -> None:
        """
        Reset reader to initial state, with first worksheet active.
        """
        self.set_worksheet()

    def set_worksheet(self, title: Optional[str] = None) -> None:
        """
        Set or reset the active worksheet.

        Args:
            title:
                Title of worksheet to use, defaults to the first worksheet if not
                provided.

        Raises:
            KeyError:
                If given worksheet title is not found.

        Returns:
            None
        """
        if title is None:
            title = self.get_worksheets()[0]

        # Cannot set `active` to a `ReadOnlyWorksheet`, have to do it by index
        index = self.workbook.index(self.workbook[title])
        self.workbook.active = index

    def _build_fieldnames(self, header_row: Tuple[Any, ...]) -> Tuple[str, ...]:
        """
        Create data keys from header row.

        Args:
            header_row:
                Header row from spreadsheet.

        Returns:
            Clean and unique strings.
        """
        cleaned = rstrip_iterable(header_row)
        cleaned = ['' if v is None else v for v in cleaned]
        fieldnames = make_slugs(cleaned)
        return tuple(fieldnames)

    def _find_header(
        self,
        rows: Iterable[Any],
        *,
        min_values=3,
        abort_after=10,
    ) -> Tuple[Any, ...]:
        """
        Consume rows from given generator until header is found.

        The first row containing more than `min_values` is considered to
        be the header. While simple, this heuristic neatly skips headings,
        dates, etc.

        Args:
            rows:
                Generator from workbook's `iter_rows()` method.
            min_values:
                Override the minimum number of rows needed in header.
            abort_after:
                Overide the maximum number of rows examined.

        Raises:
            ExcelReaderError:
                If no header could be found before `abort_after` rows
                were examined.

        Return:
            First row that looks like a header.
        """
        headers: Optional[Tuple[Any, ...]] = None
        num_searched = 0
        for index, row in enumerate(rows, 1):
            if count_values(row) >= min_values:
                headers = row
                break
            else:
                num_searched += 1

            if num_searched > abort_after:
                break

        if headers is None:
            message = f"Could not find data in worksheet {self.get_worksheet()!r}"
            raise ExcelReaderError(message)

        return headers
